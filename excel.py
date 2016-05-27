# -*- coding: utf-8 -*-
from openpyxl import load_workbook

class ExcelImport:
    
    def __init__(self, filename): #클래스 생성할 때 이름 넣어주기
        self.filename = filename
     
    def loadFile(self): #파일 불러오기
        self.wb = load_workbook(self.filename)
        sheetList = self.wb.get_sheet_names()
        self.sheet = self.wb.get_sheet_by_name(sheetList[0])
        self.maxCol = self.sheet.max_column
        self.maxRow = self.sheet.max_row;
        self.keys = []
        self.datas = []
        for x in range(1,self.sheet.max_column+1):
            self.keys.append(self.sheet.cell(row = 1, column = x).value)
        for x in range(2,self.sheet.max_row+1):
            data = []            
            for y in range(1,self.sheet.max_column+1):
                data.append(self.sheet.cell(row = x,column = y).value)
            self.datas.append(data)
    
    def printInfo(self):#정보 출력하기
        for x in range(0,self.maxRow):
            for y in range(0,self.maxCol):
                print("{0}:{1}".format(self.keys[y],self.datas[x][y]), end = ' ')
                
    def filterPrint(self,idx,Contents): #정보가 몇번째 정보인지 알고 비교할 내용을 넣어주면 자동으로 필터링하여 출력
        for x in range(0,self.maxRow):
            for y in range(0,self.maxCol):
                if self.datas[x][idx] == Contents:
                    print("{0}:{1}".format(self.keys[y],self.datas[x][y]), end = ' ')
  